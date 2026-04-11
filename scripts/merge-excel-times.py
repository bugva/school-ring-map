#!/usr/bin/env python3
"""Excel Sayfa1: mor→mavi (ilk 28), kahve→kahve, sarı→sarı, sarıkırmızı→kırmızı ring (durak no sarıdan sonra devam)."""
import json
import sys
from datetime import time as dtime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RINGS_PATH = ROOT / "public/data/rings.json"
XLSX_DEFAULT = Path.home() / "Desktop" / "ring saatleri .xlsx"


def bucket(h: int, m: int) -> str:
    hm = h * 60 + m
    if hm >= 17 * 60 or hm < 5 * 60:
        return "weekday_evening"
    return "weekday_day"


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX_DEFAULT
    if not xlsx_path.is_file():
        print(f"XLSX bulunamadı: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    data = json.loads(RINGS_PATH.read_text(encoding="utf-8"))
    by_id = {r["id"]: r for r in data["rings"]}
    mavi = by_id.get("mavi")
    kahve = by_id.get("kahve")
    sari = by_id.get("sari")
    kirmizi = by_id.get("kirmizi")
    if not kahve:
        print("kahve ring yok", file=sys.stderr)
        sys.exit(1)
    if not sari:
        print("sarı ring yok", file=sys.stderr)
        sys.exit(1)
    if not kirmizi:
        print("kirmizi ring yok", file=sys.stderr)
        sys.exit(1)
    if not mavi:
        print("Uyarı: mavi ring yok; Excel 'mor' satırları uygulanmayacak.", file=sys.stderr)

    if mavi:
        for s in mavi["stops"]:
            s["times"] = {"weekday_day": [], "weekday_evening": [], "weekend": []}
    for s in kahve["stops"]:
        s["times"] = {"weekday_day": [], "weekday_evening": [], "weekend": []}
    for s in sari["stops"]:
        s["times"] = {"weekday_day": [], "weekday_evening": [], "weekend": []}
    for s in kirmizi["stops"]:
        s["times"] = {"weekday_day": [], "weekday_evening": [], "weekend": []}

    acc_mavi: dict[str, dict[str, list[str]]] = (
        {s["id"]: {"weekday_day": [], "weekday_evening": []} for s in mavi["stops"][:28]}
        if mavi
        else {}
    )
    acc_kahve: dict[str, dict[str, list[str]]] = {
        s["id"]: {"weekday_day": [], "weekday_evening": []} for s in kahve["stops"]
    }
    acc_sari: dict[str, dict[str, list[str]]] = {
        s["id"]: {"weekday_day": [], "weekday_evening": []} for s in sari["stops"]
    }
    acc_kirmizi: dict[str, dict[str, list[str]]] = {
        s["id"]: {"weekday_day": [], "weekday_evening": []} for s in kirmizi["stops"]
    }
    n_sari = len(sari["stops"])

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Sayfa1"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        label = str(row[0]).strip().lower()
        if label == "label;time;stop":
            continue
        t = row[2]
        idx = row[3]
        if t is None or idx is None:
            continue
        if not isinstance(t, dtime):
            continue
        b = bucket(t.hour, t.minute)
        hm = f"{t.hour:02d}:{t.minute:02d}"
        if label == "mor":
            if not mavi:
                continue
            i = int(idx) - 1
            if i < 0 or i >= 28 or i >= len(mavi["stops"]):
                continue
            sid = mavi["stops"][i]["id"]
            acc_mavi.setdefault(sid, {"weekday_day": [], "weekday_evening": []})
            acc_mavi[sid][b].append(hm)
        elif label == "kahve":
            i = int(idx) - 1
            if i < 0 or i >= len(kahve["stops"]):
                continue
            sid = kahve["stops"][i]["id"]
            acc_kahve[sid][b].append(hm)
        elif label == "sarı":
            i = int(idx) - 1
            if i < 0 or i >= n_sari:
                continue
            sid = sari["stops"][i]["id"]
            acc_sari[sid][b].append(hm)
        elif label == "sarıkırmızı":
            i = int(idx) - n_sari - 1
            if i < 0 or i >= len(kirmizi["stops"]):
                continue
            sid = kirmizi["stops"][i]["id"]
            acc_kirmizi[sid][b].append(hm)

    def apply_acc(ring: dict, acc: dict) -> None:
        for s in ring["stops"]:
            sid = s["id"]
            if sid not in acc:
                continue
            for key in ("weekday_day", "weekday_evening"):
                s["times"][key] = sorted(set(acc[sid][key]))
            s["times"]["weekend"] = sorted(
                set(s["times"]["weekday_day"] + s["times"]["weekday_evening"])
            )

    if mavi:
        apply_acc(mavi, acc_mavi)
    apply_acc(kahve, acc_kahve)
    apply_acc(sari, acc_sari)
    apply_acc(kirmizi, acc_kirmizi)

    RINGS_PATH.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print("OK", RINGS_PATH)


if __name__ == "__main__":
    main()
